from django.shortcuts import render, redirect
from .models import Student, Attendance
from .forms import AttendanceForm
from django.contrib.auth.decorators import login_required
import datetime
from django.db.models import Count
from django.utils.timezone import now
import csv
from django.http import HttpResponse
from django.shortcuts import render
from calendar import monthrange
import openpyxl
from openpyxl.utils import get_column_letter

from django.shortcuts import render, redirect
from django.contrib import messages
def home(request):
    return render(request, "home.html")


def mark_attendance(request):
    today = datetime.date.today()

    if request.method == "POST":
        ATTENDANCE_DATA[today] = {}
        for student in STUDENTS:
            status = request.POST.get(student, "Absent")  # default is Absent
            ATTENDANCE_DATA[today][student] = status
        return redirect("attendance_list")

    return render(request, "mark_attendance.html", {"students": STUDENTS})


def attendance_list(request):
    # Fake attendance records: all Present (or you can add logic to vary it)
    today = datetime.date.today()
    today_data = ATTENDANCE_DATA.get(today, {})

    return render(request, "attendance_list.html", {
        "students": STUDENTS,
        "attendance": today_data,
        "date": today
    })


def weekly_summary(request):
    today = datetime.date.today()
    start_week = today - datetime.timedelta(days=today.weekday())  # Monday
    end_week = start_week + datetime.timedelta(days=6)             # Sunday

    # Fake attendance data: randomly assign Present/Absent for each day
    report = {}
    for student in STUDENTS:
        report[student] = {}
        for i in range(7):
            day = start_week + datetime.timedelta(days=i)
            # Example: mark all Present, you can change logic to fetch from DB
            report[student][day] = "Present" if i % 2 == 0 else "Absent"

    context = {
        "report": report,
        "week": f"{start_week} to {end_week}",
        "days": [(start_week + datetime.timedelta(days=i)) for i in range(7)]
    }

    return render(request, 'weekly_summary.html', context)


def download_weekly_csv(request):
    today = now().date()
    start_week = today - datetime.timedelta(days=today.weekday())   # Monday
    end_week = start_week + datetime.timedelta(days=6)              # Sunday

    summary = (
        Attendance.objects.filter(date__range=[start_week, end_week])
        .values("student__user__username", "status")
        .annotate(count=Count("id"))
    )

    # Organize data
    report = {}
    for record in summary:
        student = record["student__user__username"]
        status = record["status"]
        count = record["count"]
        if student not in report:
            report[student] = {"Present": 0, "Absent": 0}
        report[student][status] = count

    # Create CSV response
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="weekly_summary_{start_week}_to_{end_week}.csv"'

    writer = csv.writer(response)
    writer.writerow(["Student", "Present", "Absent"])
    for student, data in report.items():
        writer.writerow([student, data["Present"], data["Absent"]])

    return response


def download_weekly_excel(request):
    today = now().date()
    start_week = today - datetime.timedelta(days=today.weekday())
    end_week = start_week + datetime.timedelta(days=6)

    summary = (
        Attendance.objects.filter(date__range=[start_week, end_week])
        .values("student__user__username", "status")
        .annotate(count=Count("id"))
    )

    report = {}
    for record in summary:
        student = record["student__user__username"]
        status = record["status"]
        count = record["count"]
        if student not in report:
            report[student] = {"Present": 0, "Absent": 0}
        report[student][status] = count

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Weekly Attendance"

    # Header
    ws.append(["Student", "Present", "Absent"])

    # Data
    for student, data in report.items():
        ws.append([student, data["Present"], data["Absent"]])

    # Adjust column width
    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    # Response
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="weekly_summary_{start_week}_to_{end_week}.xlsx"'
    wb.save(response)
    return response



def monthly_summary(request):
    today = datetime.date.today()
    start_month = today.replace(day=1)
    last_day = monthrange(today.year, today.month)[1]
    end_month = today.replace(day=last_day)

    # Fake attendance data: alternate Present/Absent for demo
    report = {}
    for student in STUDENTS:
        report[student] = {}
        for day_num in range(1, last_day + 1):
            day = today.replace(day=day_num)
            report[student][day] = "Present" if day_num % 2 == 0 else "Absent"

    context = {
        "report": report,
        "month": today.strftime("%B %Y"),
        "days": [today.replace(day=i) for i in range(1, last_day + 1)]
    }

    return render(request, "monthly_summary.html", context)


def download_monthly_csv(request):
    today = now().date()
    start_month = today.replace(day=1)
    last_day = monthrange(today.year, today.month)[1]
    end_month = today.replace(day=last_day)

    summary = (
        Attendance.objects.filter(date__range=[start_month, end_month])
        .values("student__user__username", "status")
        .annotate(count=Count("id"))
    )

    report = {}
    for record in summary:
        student = record["student__user__username"]
        status = record["status"]
        count = record["count"]
        if student not in report:
            report[student] = {"Present": 0, "Absent": 0}
        report[student][status] = count

    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="monthly_summary_{start_month}_to_{end_month}.csv"'

    writer = csv.writer(response)
    writer.writerow(["Student", "Present", "Absent"])
    for student, data in report.items():
        writer.writerow([student, data["Present"], data["Absent"]])

    return response



def download_monthly_excel(request):
    today = now().date()
    start_month = today.replace(day=1)
    last_day = monthrange(today.year, today.month)[1]
    end_month = today.replace(day=last_day)

    summary = (
        Attendance.objects.filter(date__range=[start_month, end_month])
        .values("student__user__username", "status")
        .annotate(count=Count("id"))
    )

    report = {}
    for record in summary:
        student = record["student__user__username"]
        status = record["status"]
        count = record["count"]
        if student not in report:
            report[student] = {"Present": 0, "Absent": 0}
        report[student][status] = count

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Monthly Attendance"

    ws.append(["Student", "Present", "Absent"])
    for student, data in report.items():
        ws.append([student, data["Present"], data["Absent"]])

    for col in ws.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[get_column_letter(column)].width = max_length + 2

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="monthly_summary_{start_month}_to_{end_month}.xlsx"'
    wb.save(response)
    return response



# Example students list
STUDENTS = [
    "Narisetty Harshitha","Tirumala Keerthi","K Yoga Nandini Reddy","Ramavath Sirisha Bai","Balla Prathima","Y.Lakshmi Prathyusha"
    "M.Pujitha Devi","Appana Nikhitha","Kanumuri Chinmayee","Kantheti Latha","Kondeti Hema Tulasi","U.poorna Sai Bhavana","Rajana Sravika"
    ,"D HemaLatha","Gajjala Iswarya","Obenaboina Lavnaya","Prasanthi Kamanuru","Akshaya Deepika Eerla ","Hemalatha Vennapusa","R.Chakra Varshini",
    "Puli Sreehitha Reddy","M Jayasree","NV Sravani","Hemambika Kouthavarapu","D Sandhya Rani"
]
ATTENDANCE_DATA = {}